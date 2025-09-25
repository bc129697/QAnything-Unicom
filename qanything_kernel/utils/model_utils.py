import os
import tiktoken
import numpy as np
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from qanything_kernel.utils.custom_log import debug_logger
from qanything_kernel.configs.model_config import UPLOAD_ROOT_PATH


# from transformers import AutoTokenizer
# embedding_tokenizer = AutoTokenizer.from_pretrained(LOCAL_EMBED_PATH, local_files_only=True)
# rerank_tokenizer = AutoTokenizer.from_pretrained(LOCAL_RERANK_PATH, local_files_only=True)
encoding = tiktoken.encoding_for_model('gpt-3.5-turbo-0613')
# encoding = tiktoken.encoding_for_model('cl100k_base')


def num_tokens(text: str) -> int:
    """Return the number of tokens in a string."""
    return len(encoding.encode(text, disallowed_special=()))


def num_tokens_embed(text: str) -> int:
    """Return the number of tokens in a string."""
    # return len(embedding_tokenizer.encode(text, add_special_tokens=True))
    return num_tokens(text)


def num_tokens_rerank(text: str) -> int:
    """Return the number of tokens in a string."""
    # return len(rerank_tokenizer.encode(text, add_special_tokens=True))
    return num_tokens(text)


def num_tokens_from_messages(message_texts):
    num_tokens = 0
    for message in message_texts:
        # num_tokens += 4  # every message follows <im_start>{role/name}\n{content}<im_end>\n
        # for key, value in message.items():
        num_tokens += len(encoding.encode(message, disallowed_special=()))
        # if key == "name":  # if there's a name, the role is omitted
        # num_tokens += -1  # role is always required and always 1 token
    # num_tokens += 2  # every reply is primed with <im_start>assistant
    return num_tokens

def export_qalogs_to_excel(qalogs, columns, filename: str):
    # 将查询结果转换为 DataFrame
    df = pd.DataFrame(qalogs, columns=columns)

    # 写入 Excel 文件
    root_path = os.path.dirname(UPLOAD_ROOT_PATH) + '/saved_qalogs'
    if not os.path.exists(root_path):
        os.makedirs(root_path)

    file_path = os.path.join(root_path, filename)
    df.to_excel(file_path, index=False)

    # 使用 openpyxl 调整列宽
    workbook = load_workbook(filename=file_path)
    worksheet = workbook.active

    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

    workbook.save(file_path)
    return file_path


def cosine_similarity(embedding1, embedding2):
    embedding1 = np.array(embedding1)
    embedding2 = np.array(embedding2)
    # 计算两个向量的点积
    dot_product = np.dot(embedding1, embedding2)
    # 计算两个向量的模
    norm_embedding1 = np.linalg.norm(embedding1)
    norm_embedding2 = np.linalg.norm(embedding2)
    # 计算余弦相似度
    similarity = dot_product / (norm_embedding1 * norm_embedding2)
    # 将余弦相似度映射到0-1之间
    similarity_mapped = (similarity + 1) / 2
    return similarity_mapped


def check_and_transform_excel(binary_data):
    # 使用BytesIO读取二进制数据
    try:
        data_io = BytesIO(binary_data)
        df = pd.read_excel(data_io)
    
        # 检查列数
        if len(df.columns) != 2:
            return "格式错误：文件应该只有两列"

        # 检查列标题
        if df.columns[0] != "问题" or df.columns[1] != "答案":
            return "格式错误：第一列标题应为'问题'，第二列标题应为'答案'"

        # 检查每行长度
        for index, row in df.iterrows():
            question_len = len(str(row['问题']))
            answer_len = len(str(row['答案']))
            if question_len > 512 or answer_len > 8000:
                return f"行{index + 1}长度超出限制：问题长度={question_len}，答案长度={answer_len}"

        # 转换数据格式
        transformed_data = []
        for _, row in df.iterrows():
            transformed_data.append({"question": str(row['问题']), "answer": str(row['答案'])})

        return transformed_data
    except Exception as e:
        debug_logger.error(f"Error processing Excel file: {str(e)}")
        return f"读取文件时出错: {e}"


def get_vllm_model_length(model_url, model_name):
    """
    获取vLLM模型的最大长度限制
    
    Args:
        model_url (str): 模型服务的URL地址
        model_name (str): 模型名称
    
    Returns:
        int: 模型的最大长度限制
    """
    import requests
    from qanything_kernel.configs.model_config import LOCAL_RERANK_MAX_LENGTH
    
    try:
        # 移除URL中的v1部分以获取基础URL
        base_url = model_url.split("v1")[0]
        
        # 构造获取模型信息的URL
        url = f"{base_url}v1/models"
        
        # 发送请求获取模型信息
        response = requests.get(url, timeout=5)
        response.raise_for_status()
        
        # 解析响应获取模型最大长度
        model_info = response.json()
        
        # 根据返回的数据结构解析max_model_len
        # 数据结构: {"object": "list", "data": [{"id": "...", "max_model_len": 2048, ...}]}
        if 'data' in model_info and isinstance(model_info['data'], list):
            for model_data in model_info['data']:
                if model_data.get('id') == model_name and 'max_model_len' in model_data:
                    return model_data['max_model_len']
        
        # 如果无法从模型信息获取，则返回默认值
        return LOCAL_RERANK_MAX_LENGTH
        
    except Exception as e:
        # 如果请求失败或解析出错，返回默认值
        return LOCAL_RERANK_MAX_LENGTH